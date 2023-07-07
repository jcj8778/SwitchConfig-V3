import openpyxl

# file = input("File path: ")
# file_path = 'C:/Users/jjones/Downloads/Fishers Newell OT Network Tool v1(2) copy.xlsx'
file_path = 'C:/Users/jjones/Documents/Fishers Newell OT Network Tool v1(2) copy.xlsx'
# Load the Excel workbook
workbook = openpyxl.load_workbook(file_path)
wf_workbook = openpyxl.Workbook()
ncc_sheets = [sheet for sheet in workbook.sheetnames if sheet.startswith('NCC')]

for sheet_name in ncc_sheets:
    sheet = workbook[sheet_name]

    # Create a new sheet in the condensed workbook
    condensed_sheet = wf_workbook.create_sheet(title=sheet_name)

    # Initialize previous name variable
    prev_name = None

    # Iterate over each row in the sheet
    for row in sheet.iter_rows(values_only=True):
        condensed_row = [row[2], row[3], row[7]]
        name = row[1]
        if name in ['Switch']:
            continue
        if name is not None and name != prev_name:
            prev_name = name
        if any(cell is None or cell == '' for cell in condensed_row):
            continue
        condensed_sheet.append([prev_name] + condensed_row)

# Remove the default sheet created by openpyxl
wf_workbook.remove(wf_workbook["Sheet"])

# Save the condensed workbook
wf_workbook.save('condensed_workbook.xlsx')

# Close the workbooks
workbook.close()
wf_workbook.close()
