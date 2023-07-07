import openpyxl
import switch as switch
from netmiko import ConnectHandler

from Switch import Switch

file = 'condensed_workbook.xlsx'
workbook = openpyxl.load_workbook(file)
switches = []

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows(values_only=True):
        switch_name = row[0]
        port = row[1]
        description = row[2]
        vlan_id = row[3]

        switches.append(switch)
        if switch_name not in switches:
            switches[switch_name] = Switch(switch_name)

        switches[switch_name].add_port(port, description, vlan_id)

print(switches)