import openpyxl
from netmiko import ConnectHandler

from Switch import Switch

file = 'condensed_workbook.xlsx'
workbook = openpyxl.load_workbook(file)
switches = []

for sheet in workbook.sheetnames:
    for row in sheet.iter_rows(values_only=True):
        switch_name = 'FSH-' + row[0]
        port = row[1]
        description = row[2]
        vlan_id = row[3]
        print(switch.get_name())
        switches.append(switch)
        if switch_name not in switches:
            switches[switch_name] = Switch(switch_name)

        switches[switch_name].add_port()
    print(switches)
    tokens = [29, 13, 19, 24]
    for token in tokens:
        ip = "172.16.150." + str(token)
        # SwitchName = input("Switch Name: ")
        # Define the device information
        device = {
            'device_type': 'cisco_ios',
            'ip': ip,
            'username': 'admin',
            'password': 'switch',
        }


            expected_values = {
                'Switch hostname': host,
                'IP IGMP Snooping': 'IGMP snooping                : Enabled',
                'REP Admin vlan': 'Admin-vlan: 600',
                'Port Fast and bpdufilter': 'spanning-tree portfast bpdufilter default',
                'PVST': 'spanning-tree mode rapid-pvst',
                'alarm-profile': 'default',
                # 'REP Interfaces': 'Correct Descriptions, native vlan 600, mode trunk, seg 1',
                # 'Trunk Interfaces': 'Correct Description, native vlan 600, mode trunk, trunk allowed vlans 600-603,614-618',
                # 'Disabled Ports': 'Disabled (Description: Disabled Admin Down, access vlan 1, shutdown)',
                # 'Gi2/16': 'Maintenance Port, access vlan, mode access',
                # 'App Interface': 'shutdown',
                # 'interface Vlan1': 'no ip address',
                # 'Management Interface': 'Management interface vlan (600, 10.192.202.13, 255.255.255.0, No IPv6)',
                'CIP enabled': 'State : Enabled',
                'Gateway': '10.192.202.1',
                'NO HTTP': 'no ip http server',
                'http secure-server': 'ip http secure-server',
                'Logging Host': '10.192.202.6',
                'Message of the Day': 'WARNING:',
                'Logon Message': 'WARNING: No privacy expectation. For official use only',
                'Firmware version': 'System image file is "sdflash:/s5800-universalk9.17.10.01.SPA.bin"',
                # 'REP Ports': 'GI1/1&GI1/2 set to trunk and REP enabled',
                # 'REP Port Type': 'Transit',
                # 'REP Health': 'Verified REP partners and status',
            }

            # Establish SSH connection to the switch
            net_connect = ConnectHandler(**device)

            # Define the commands to check for each setting
            commands = {
                'Switch hostname': 'show run | include hostname',
                'IP IGMP Snooping': 'show ip igmp snooping',
                'REP Admin vlan': 'show interfaces rep detail',
                'Port Fast and bpdufilter': 'show running-config | include spanning-tree',
                'PVST': 'show running-config | include spanning-tree',
                'alarm-profile': 'show alarm profile',
                # 'Interface description': 'show interface description',
                # 'Trunk Interfaces': 'show interface trunk',
                # 'Disabled Ports': 'show interface status | include Disabled',
                # 'Gi2/16': 'show interface gigabitEthernet 2/16',
                # 'App Interface': 'show interface AppGigabitEthernet1/1',
                # 'interface Vlan1': 'show interface vlan 1',
                # 'Management Interface': 'show interface vlan 600',
                'CIP enabled': 'show cip status',
                'Gateway': 'show run | include default-gateway',
                'NO HTTP': 'show run | include ip http server',
                'http secure-server': 'show run | include ip http secure-server',
                'Logging Host': 'show running-config | include logging',
                'Message of the Day': 'show banner motd',
                'Logon Message': 'show banner login',
                'Firmware version': 'show version | include bin',
                # 'REP Ports': 'show rep port',
                # 'REP Port Type': 'show rep port-type',
                # 'REP Health': 'show rep health',
            }
            print (switch)
            # Compare the expected values with the actual output
            for setting, command in commands.items():
                output = net_connect.send_command(command)

                # Compare the expected value with the actual output
                if setting in expected_values and expected_values[setting] in output:
                    print(f'{setting}: Passed')

                else:
                    print(f'{setting}: Failed')
                    print(f'Expected: {expected_values[setting]}')
                    print(f'Actual: {output}')
                print("=" * 50)

            # Disconnect from the switch

            net_connect.disconnect()
